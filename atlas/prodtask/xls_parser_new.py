'''
Created on Feb 5, 2014

@author: mborodin
'''
import tempfile
import xml.dom.minidom
import zipfile

import requests
import xlrd
import openpyxl
def open_tempfile_from_url(url,tempfile_suffix):
        filename = tempfile.mkstemp(suffix=tempfile_suffix)[1]
        response = requests.get(url)
        response_status_code = response.status_code
        if (response.status_code != 200):
            raise RuntimeError("URL %s return code %i" % (url, response_status_code))
            
        with open(filename, 'wb') as output:
            for buf in response.iter_content(65536):
                output.write(buf)
        return filename

def open_google_ss(google_key, output_format):
        """Download google spread sheet to a local temp file. 

        :param google_key: A key for google sreapsheet.
        :param output_format: supported google download format(csv,xls,ods)

        Returns a temp file path.
        """
        url = "https://docs.google.com/spreadsheet/ccc?output=%s&key=%s" % (output_format, google_key)
        return open_tempfile_from_url(url,'.%s' % output_format)

class XlrParser(object):
     
    
    def __init__(self):
        self.result_dictr = {}
        
    def open_by_key(self,google_key, format):
        """Parse google spreadsheet as xls file. 

        :param google_key: A key for google sreapsheet.


        Returns a dict tuple (result dict, color dict)
        """
        with open(open_google_ss(google_key,format),'rb') as f:
            res = self.__parse_xsl_openpyxl(f)
        return res
            
    def open_by_open_file(self,file_object):
        """Parse   xls file. 

        :param file_object: An open file descriptor.


        Returns a dict tuple (result dict, color dict)
        """
        return self.__parse_xsl_openpyxl(file_object)
    
    def __parse_xsl(self, file_object):     
        nocolor = False
        file_contents = file_object.read()
        try:
            book = xlrd.open_workbook(file_contents=file_contents, formatting_info=1)
        except NotImplementedError:
            book = xlrd.open_workbook(file_contents=file_contents)
            nocolor = True
        self.result_dict = {}
        color_dict = {}
        slice_index = 2
        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            for row_index in range(sheet.nrows):
                row_dict = {}
                row_color_dict = {}
                for col_index in range(sheet.ncols):
                    if(row_index >= 1):
                        cell_value = sheet.cell(row_index, col_index).value
                        if cell_value:
                            row_dict[col_index] = cell_value
                            if not nocolor:
                                xfx = sheet.cell_xf_index(row_index, col_index)
                                xf = book.xf_list[xfx]
                                row_color_dict[col_index] = xf.background.pattern_colour_index
                if row_dict:
                    self.result_dict[slice_index] = row_dict
                    
                    if color_dict:
                        color_dict[slice_index] = row_color_dict
                    slice_index += 1
        return  (self.result_dict, color_dict)

    def __parse_xsl_openpyxl(self, file_object):
        book =  openpyxl.load_workbook(file_object)
        self.result_dict = {}
        color_dict = {}
        slice_index = 2
        for sheet in book.worksheets:
            for row in sheet.iter_rows(min_row=2):
                row_dict = {}
                for cell in row:
                    if cell.value:
                        row_dict[cell.column-1] = cell.value
                if row_dict:
                    self.result_dict[slice_index] = row_dict
                    slice_index += 1
        return (self.result_dict, color_dict)
  
class OdfReader(object):
    def __init__(self, googlekey):
        """
        Open an ODF file.
        """

        self.filename = open_google_ss(googlekey, 'ods')
        self.m_odf = zipfile.ZipFile(self.filename)
        self.filelist = self.m_odf.infolist()
        self.parse_details()         # Get the rows

    def parse_details(self):
        """
        Do the thing.
        """
        ostr = self.m_odf.read('content.xml')
        doc = xml.dom.minidom.parseString(ostr)
        tbls = doc.getElementsByTagName('table:table')

        self.details = {}
        for t in tbls:
            table_name = t.attributes['table:name'].value
            # Select first spreadsheet
            rows = t.getElementsByTagName('table:table-row')
            nrows = 1
            for r in rows:
                cells = r.getElementsByTagName('table:table-cell')
                htxt = {}
                ic = 0
                #alab=["brief","ds","format","jo","evfs","eva2","prio","evgen","simul","merge","digi","reco","rmerge","rtag","a2","a2merge","a2tag","LO","feff","NLO","gen","ecm","ef","comm","contact","store"]
                for c in cells:
                    entry = c.firstChild
                    stxt = ''
                    try:
                        rep = int(c.attributes['table:number-columns-repeated'].value)
                    except:
                        rep = 1
                    if entry != None:
                        stxt = "notext" #TODO: Why it's here?
                        sstat = ""
                        for t in entry.childNodes:
                            if t.nodeType == t.TEXT_NODE:
                                intxt = t.data.encode('utf-8')
                                if stxt.find("notext") != -1:
                                    stxt = intxt
                                else:
                                    stxt = stxt + intxt

                    if nrows > 1 and len(stxt) > 0 :
                        htxt[ic] = stxt
                    ic += rep

                if htxt:
                    self.details[nrows] = htxt
                nrows += 1

    def get_details(self):
        """
        Return dictionary with contents
        """
        return self.details
        
